"""Split a workbook into one .xlsx per sheet, for reading rather than promoting.

    python -m scripts.split_workbook output/workbooks/cbb.xlsx
    python -m scripts.split_workbook output/workbooks/cbb.xlsx --out some/dir

A `promote` reads the ONE multi-sheet workbook, because the sheets reference each
other: `regulations.compliancecategory_id` points into `compliancecategory`, and
`regulation_versions.regulation_id` into `regulations`. Split apart those links
still exist as numbers but nothing resolves them.

So these files are for a PERSON to read — one tab per file, openable side by
side, no scrolling between sheets. THE SPLIT FILES ARE NOT PROMOTABLE. Promote
the original.

OVERSIZED VALUES. Excel caps a cell at 32,767 characters and the exporter puts
anything longer in a `.fulltext.json` sidecar, leaving a marker in the cell. This
copies the marker as-is rather than rehydrating: a 200 KB document text pasted
into a cell would be unreadable anyway, and silently truncating it would be
worse. Where a sidecar exists its name is printed, so the pair stays together.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

import openpyxl                                        # noqa: E402

#: Windows forbids these in a filename, and a sheet name may hold any of them.
_UNSAFE = re.compile(r'[<>:"/\\|?*]')


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("workbook")
    ap.add_argument("-o", "--out", help="default: <workbook stem>_sheets/ beside it")
    ap.add_argument("--skip-empty", action="store_true",
                    help="do not write a file for a sheet with no data rows")
    a = ap.parse_args()

    src = Path(a.workbook)
    if not src.exists():
        raise SystemExit(f"no such workbook: {src}")
    out = Path(a.out) if a.out else src.with_name(f"{src.stem}_sheets")
    out.mkdir(parents=True, exist_ok=True)

    # read_only=True streams instead of loading the whole book, which matters on
    # the big ones -- CBB's processing_log alone is 4,035 rows.
    wb = openpyxl.load_workbook(src, read_only=True)
    print(f"{src}  ->  {out}\n")

    written = 0
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = max(len(rows) - 1, 0)
        if a.skip_empty and not data_rows:
            print(f"  {name:24} {data_rows:>6} rows   skipped (empty)")
            continue

        dst_wb = openpyxl.Workbook()
        dst = dst_wb.active
        dst.title = name[:31]                     # Excel's own limit
        for row in rows:
            dst.append(list(row))
        if rows:
            dst.freeze_panes = "A2"               # keep the header visible
        safe = _UNSAFE.sub("_", name)
        dst_path = out / f"{src.stem}__{safe}.xlsx"
        dst_wb.save(dst_path)
        written += 1
        print(f"  {name:24} {data_rows:>6} rows   {dst_path.name}")

    sidecar = src.with_suffix(".fulltext.json")
    print()
    if sidecar.exists():
        print(f"NOTE: {sidecar.name} holds this workbook's oversized cell values. "
              f"The split files carry the MARKERS, not the text.")
    print(f"{written} file(s) written to {out}")
    print("These are for reading. Promote the original workbook, not these — "
          "the sheets reference each other by id and a single sheet cannot "
          "resolve those on its own.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
