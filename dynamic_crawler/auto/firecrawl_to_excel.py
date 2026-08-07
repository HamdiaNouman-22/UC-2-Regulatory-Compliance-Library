"""Firecrawl scrape-based tree-walk -> Excel.  Purely additive, local-only.

Plain-English purpose: start at ONE regulator seed page, use Firecrawl `scrape`
(which renders JavaScript, so it actually sees the links on JS-heavy sites like
the SAMA rulebook), follow those links down the tree, and dump EVERYTHING that
came back for each page into an Excel file you can open and eyeball.

Why scrape and not crawl: Firecrawl's `crawl` endpoint returns 0 on this site
because its link-discovery doesn't render JavaScript. `scrape` does. So we walk
the tree ourselves using the links scrape hands back. (See firecrawl_crawl_test.py.)

Never touches the live pipeline or the production DB. Writes only under
output/firecrawl_test/<label>/.

Usage:
    python -m dynamic_crawler.auto.firecrawl_to_excel
    python -m dynamic_crawler.auto.firecrawl_to_excel --url <seed> --max-pages 20 --max-depth 3
"""

import argparse
import json
import logging
import os
import re
import time
from collections import Counter, deque
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

API_BASE = os.getenv("FIRECRAWL_API_URL", "https://api.firecrawl.dev/v2").rstrip("/")
DEFAULT_SEED = "https://rulebook.sama.gov.sa/en/book-category/1365"
OUTPUT_DIR = Path("output") / "firecrawl_test"

# Free plan is throttled (~3 req/min). Space requests out and back off on 429.
MIN_INTERVAL_SECONDS = 22
EXCEL_CELL_LIMIT = 32000  # Excel hard limit is 32767; leave headroom.


class FirecrawlError(RuntimeError):
    pass


def _headers(api_key: str) -> dict:
    return {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}


class RateLimitedScraper:
    def __init__(self, api_key: str):
        self.h = _headers(api_key)
        self._last = 0.0

    def scrape(self, url: str) -> dict:
        wait = MIN_INTERVAL_SECONDS - (time.monotonic() - self._last)
        if wait > 0:
            time.sleep(wait)
        body = {"url": url, "formats": ["markdown", "links"]}
        last_err = "unknown"
        for attempt in range(4):
            try:
                r = requests.post(f"{API_BASE}/scrape", headers=self.h, json=body, timeout=120)
            except requests.exceptions.RequestException as e:
                # Network blip (timeout / DNS / reset): back off and retry, don't crash.
                last_err = f"network error: {type(e).__name__}"
                logger.warning("network error on %s (attempt %d/4): %s", url, attempt + 1, e)
                time.sleep(10 * (attempt + 1))
                continue
            finally:
                self._last = time.monotonic()
            if r.status_code == 429:
                logger.warning("rate limited on %s; waiting 40s", url)
                time.sleep(40)
                continue
            if r.status_code >= 400:
                return {"error": f"HTTP {r.status_code}: {r.text[:300]}"}
            return r.json().get("data", {}) or {}
        return {"error": last_err}


def same_site_doc_links(links, seed_host: str):
    """Keep on-domain /en/ links, drop anchors/assets, dedupe, preserve order."""
    out, seen = [], set()
    for u in links or []:
        if not u or "#" in u:
            continue
        p = urlparse(u)
        if p.netloc and p.netloc != seed_host:
            continue
        if "/en/" not in p.path:
            continue
        if re.search(r"\.(pdf|jpg|jpeg|png|gif|css|js|svg)$", p.path, re.I):
            continue
        clean = u.split("#")[0]
        if clean not in seen:
            seen.add(clean)
            out.append(clean)
    return out


def flatten_metadata(meta: dict) -> dict:
    """Turn nested/list metadata values into flat, Excel-friendly scalars."""
    flat = {}
    for k, v in (meta or {}).items():
        if isinstance(v, (list, tuple)):
            v = " | ".join(str(x) for x in v)
        elif isinstance(v, dict):
            v = json.dumps(v, ensure_ascii=False)
        flat[f"meta_{k}"] = v
    return flat


def clip(s, limit=EXCEL_CELL_LIMIT):
    if isinstance(s, str) and len(s) > limit:
        return s[:limit] + f"\n...[truncated {len(s) - limit} chars — full text in pages_raw.json]"
    return s


# ---- Excel formatting -------------------------------------------------------

CORE_COLS = ["depth", "url", "parent_url", "error", "title", "description", "status_code",
             "num_links", "num_child_links", "markdown_chars", "child_links", "all_links", "markdown"]

COL_WIDTHS = {
    "depth": 6, "url": 55, "parent_url": 55, "error": 22, "title": 42, "description": 50,
    "status_code": 10, "num_links": 9, "num_child_links": 11, "markdown_chars": 12,
    "child_links": 60, "all_links": 60, "markdown": 90,
}
WRAP_COLS = {"description", "child_links", "all_links"}   # lists/short text: safe to wrap
LINK_COLS = {"url", "parent_url"}                          # rendered as clickable hyperlinks

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_LINK_FONT = Font(color="0563C1", underline="single")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_sheet(ws, df, freeze=True):
    """Apply header style, widths, wrap, borders, filter, freeze to a worksheet."""
    ncols = len(df.columns)
    for j, col in enumerate(df.columns, start=1):
        letter = get_column_letter(j)
        ws.column_dimensions[letter].width = COL_WIDTHS.get(col, 18)
        hc = ws.cell(row=1, column=j)
        hc.fill, hc.font = _HEADER_FILL, _HEADER_FONT
        hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        hc.border = _BORDER
        wrap = col in WRAP_COLS
        for i in range(2, len(df) + 2):
            c = ws.cell(row=i, column=j)
            c.alignment = Alignment(vertical="top", wrap_text=wrap)
            c.border = _BORDER
            if col in LINK_COLS and isinstance(c.value, str) and c.value.startswith("http"):
                c.hyperlink, c.font = c.value, _LINK_FONT
    ws.row_dimensions[1].height = 28
    if freeze:
        ws.freeze_panes = "A2"
    if ncols:
        ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{len(df) + 1}"


def _safe_path(xlsx_path: Path) -> Path:
    """If the target .xlsx is open/locked in Excel, pick a free numbered name."""
    try:
        if xlsx_path.exists():
            with open(xlsx_path, "a"):
                pass
        return xlsx_path
    except PermissionError:
        for n in range(1, 100):
            alt = xlsx_path.with_name(f"{xlsx_path.stem}_{n}{xlsx_path.suffix}")
            if not alt.exists():
                logger.warning("%s is open in Excel — saving to %s instead", xlsx_path.name, alt.name)
                return alt
        raise


def write_excel(rows, edges, xlsx_path):
    xlsx_path = _safe_path(Path(xlsx_path))
    df = pd.DataFrame(rows)
    meta_cols = sorted(c for c in df.columns if c.startswith("meta_"))
    df = df[[c for c in CORE_COLS if c in df.columns] + meta_cols]
    edf = pd.DataFrame(edges) if edges else pd.DataFrame(columns=["parent", "child"])

    # Summary sheet — a quick at-a-glance readout.
    depth_counts = Counter(r["depth"] for r in rows)
    http_counts = Counter(str(r.get("status_code", "")) for r in rows)
    errs = sum(1 for r in rows if r.get("error"))
    summary = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Pages scraped", len(rows)),
        ("Pages with errors", errs),
        ("Link relationships", len(edges)),
        ("", ""),
        ("Pages by depth", ""),
        *[(f"  depth {d}", depth_counts[d]) for d in sorted(depth_counts)],
        ("", ""),
        ("Pages by HTTP status", ""),
        *[(f"  {code}", http_counts[code]) for code in sorted(http_counts)],
    ]
    sdf = pd.DataFrame(summary, columns=["metric", "value"])

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as xl:
        sdf.to_excel(xl, sheet_name="summary", index=False)
        df.to_excel(xl, sheet_name="pages", index=False)
        edf.to_excel(xl, sheet_name="link_graph", index=False)
        wb = xl.book
        wb["summary"].column_dimensions["A"].width = 26
        wb["summary"].column_dimensions["B"].width = 24
        for j in (1, 2):
            hc = wb["summary"].cell(row=1, column=j)
            hc.fill, hc.font = _HEADER_FILL, _HEADER_FONT
        wb["summary"].freeze_panes = "A2"
        _style_sheet(wb["pages"], df)
        _style_sheet(wb["link_graph"], edf)
    return df, xlsx_path


def rows_from_raw(raw_pages, seed_host):
    """Rebuild table rows + edges from a saved pages_raw.json (no API calls)."""
    rows, edges = [], []
    for rp in raw_pages:
        url, depth, parent = rp["url"], rp["depth"], rp.get("parent", "")
        data = rp.get("data") or {}
        meta = data.get("metadata") or {}
        links = data.get("links") or []
        markdown = data.get("markdown") or ""
        children = same_site_doc_links(links, seed_host)
        row = {
            "depth": depth, "url": url, "parent_url": parent, "error": data.get("error", ""),
            "title": meta.get("title", ""), "description": meta.get("description", ""),
            "status_code": meta.get("statusCode", ""), "num_links": len(links),
            "num_child_links": len(children), "markdown_chars": len(markdown),
            "child_links": clip("\n".join(children)), "all_links": clip("\n".join(links)),
            "markdown": clip(markdown),
        }
        row.update(flatten_metadata(meta))
        rows.append(row)
        for c in children:
            edges.append({"parent": url, "child": c})
    return rows, edges


def main() -> None:
    ap = argparse.ArgumentParser(description="Firecrawl scrape tree-walk -> Excel (additive, local)")
    ap.add_argument("--url", default=DEFAULT_SEED)
    ap.add_argument("--max-pages", type=int, default=15, help="Total pages to scrape (rate-limited)")
    ap.add_argument("--max-depth", type=int, default=3, help="How deep to follow the tree")
    ap.add_argument("--label", default=None)
    ap.add_argument("--from-raw", action="store_true",
                    help="Skip the API; rebuild the formatted Excel from an existing pages_raw.json")
    args = ap.parse_args()

    parsed = urlparse(args.url)
    seed_host = parsed.netloc
    # Namespace per host+path so different tabs on the same site don't overwrite
    # each other (e.g. /book-category/1365 vs /regulatory-sandbox).
    label = args.label or re.sub(r"[^a-zA-Z0-9]+", "_", f"{seed_host}_{parsed.path}").strip("_")[:80]
    out_dir = OUTPUT_DIR / label
    out_dir.mkdir(parents=True, exist_ok=True)

    # Fast path: reformat previously-saved data into a pretty Excel, no crawling.
    if args.from_raw:
        raw_path = out_dir / "pages_raw.json"
        if not raw_path.exists():
            raise FirecrawlError(f"No saved data to reformat: {raw_path} (run a crawl first)")
        raw_pages = json.loads(raw_path.read_text(encoding="utf-8"))
        rows, edges = rows_from_raw(raw_pages, seed_host)
        df, saved = write_excel(rows, edges, out_dir / "firecrawl_results.xlsx")
        print(f"\nReformatted {len(rows)} pages from {raw_path.name}")
        print(f"Excel : {saved}  ({len(df.columns)} columns)")
        return

    api_key = os.getenv("FIRECRAWL_KEY")
    if not api_key:
        raise FirecrawlError("Missing FIRECRAWL_KEY in environment / .env")

    est_min = args.max_pages * MIN_INTERVAL_SECONDS / 60
    logger.info("Seed: %s", args.url)
    logger.info("max_pages=%s max_depth=%s  (~%.1f min at 3 req/min)  -> %s",
                args.max_pages, args.max_depth, est_min, out_dir)

    scraper = RateLimitedScraper(api_key)
    queue = deque([(args.url, 0, "")])       # (url, depth, parent)
    visited = set()
    raw_pages = []
    raw_path = out_dir / "pages_raw.json"
    interrupted = False

    # Any crash or Ctrl+C still leaves whatever we've fetched: raw JSON is
    # flushed after every page, and the Excel is built from it in `finally`.
    try:
        while queue and len(raw_pages) < args.max_pages:
            url, depth, parent = queue.popleft()
            if url in visited:
                continue
            visited.add(url)

            logger.info("[%d/%d] depth=%d  %s", len(raw_pages) + 1, args.max_pages, depth, url)
            data = scraper.scrape(url)
            raw_pages.append({"url": url, "depth": depth, "parent": parent, "data": data})
            raw_path.write_text(json.dumps(raw_pages, indent=2, ensure_ascii=False), encoding="utf-8")

            if data.get("error"):
                logger.warning("  page errored, skipping its children: %s", data["error"])
                continue
            children = same_site_doc_links(data.get("links") or [], seed_host)
            for c in children:
                if depth + 1 <= args.max_depth and c not in visited:
                    queue.append((c, depth + 1, url))
    except KeyboardInterrupt:
        interrupted = True
        logger.warning("Interrupted by user — saving %d pages fetched so far.", len(raw_pages))
    except Exception as e:  # noqa: BLE001 — never lose a long rate-limited run
        logger.error("Run aborted (%s) — saving %d pages fetched so far.", e, len(raw_pages))
    finally:
        if raw_pages:
            rows, edges = rows_from_raw(raw_pages, seed_host)
            df, xlsx_path = write_excel(rows, edges, out_dir / "firecrawl_results.xlsx")
            n_err = sum(1 for r in rows if r.get("error"))
            print("\n" + "=" * 70)
            status = "INTERRUPTED" if interrupted else "finished"
            print(f"Run {status}: {len(rows)} pages saved ({n_err} errored).  Columns: {len(df.columns)}")
            print(f"Excel : {xlsx_path}")
            print(f"Raw   : {raw_path}")
            print("=" * 70)
        else:
            print("No pages fetched — nothing to save.")


if __name__ == "__main__":
    main()
