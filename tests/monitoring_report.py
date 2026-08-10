"""
Monitoring Report — read-only, no DB writes.

Produces monitoring_report.xlsx with two sheets:
  Sheet 1 "Monitoring Fetched"  — every regulation the monitoring crawler touched
                                   since the initial crawl ended (> 2026-05-28):
                                   either newly added OR got a new version row.
  Sheet 2 "Version History"     — ALL version rows for each of those regulations,
                                   oldest to newest, with HTML included.
"""

import os, sys
from pathlib import Path
from dotenv import load_dotenv
import pyodbc
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))
load_dotenv(PROJECT_ROOT / ".env")

CUTOFF     = "2026-05-28"          # last day of initial crawl
OUTPUT     = PROJECT_ROOT / "tests" / "monitoring_report_jul6.xlsx"

# Excel cell hard limit
EXCEL_MAX  = 32_000

# ── DB ────────────────────────────────────────────────────────────────────────
def conn():
    s = (
        f"DRIVER={os.getenv('MSSQL_DRIVER')};"
        f"SERVER={os.getenv('MSSQL_SERVER')};"
        f"DATABASE={os.getenv('MSSQL_DATABASE')};"
        f"UID={os.getenv('MSSQL_USERNAME')};"
        f"PWD={os.getenv('MSSQL_PASSWORD')};"
        f"TrustServerCertificate=yes;"
    )
    return pyodbc.connect(s)


# ── SQL ───────────────────────────────────────────────────────────────────────

# Sheet 1: one row per regulation that monitoring touched
SHEET1_SQL = f"""
WITH monitoring_regs AS (
    -- regs added after initial crawl
    SELECT DISTINCT r.id
    FROM regulations r
    WHERE r.regulator = 'Central Bank of Bahrain'
      AND CONVERT(date, r.created_at) > '{CUTOFF}'

    UNION

    -- existing regs that received a new version after initial crawl
    SELECT DISTINCT rv.regulation_id
    FROM regulation_versions rv
    JOIN regulations r ON r.id = rv.regulation_id
    WHERE r.regulator = 'Central Bank of Bahrain'
      AND CONVERT(date, rv.created_at) > '{CUTOFF}'
)
SELECT
    r.id                                             AS reg_id,
    CASE
        WHEN CONVERT(date, r.created_at) > '{CUTOFF}' THEN 'NEW — added by monitoring'
        ELSE                                               'EXISTING — updated by monitoring'
    END                                              AS monitoring_action,
    r.title,
    r.category,
    r.status                                         AS reg_status,
    r.source_page_url,
    r.document_url,
    r.published_date,
    CONVERT(varchar(30), r.created_at, 120)          AS reg_created_at,
    CONVERT(varchar(30), r.updated_at,  120)         AS reg_updated_at,
    r.content_hash,
    (SELECT COUNT(*) FROM regulation_versions rv2
     WHERE rv2.regulation_id = r.id)                AS total_versions,
    (SELECT COUNT(*) FROM regulation_versions rv3
     WHERE rv3.regulation_id = r.id
       AND CONVERT(date, rv3.created_at) > '{CUTOFF}')
                                                     AS monitoring_versions
FROM regulations r
JOIN monitoring_regs m ON m.id = r.id
ORDER BY r.created_at DESC, r.id;
"""

# Sheet 2: one row per version for each monitoring reg
SHEET2_SQL = f"""
WITH monitoring_regs AS (
    SELECT DISTINCT r.id
    FROM regulations r
    WHERE r.regulator = 'Central Bank of Bahrain'
      AND CONVERT(date, r.created_at) > '{CUTOFF}'

    UNION

    SELECT DISTINCT rv.regulation_id
    FROM regulation_versions rv
    JOIN regulations r ON r.id = rv.regulation_id
    WHERE r.regulator = 'Central Bank of Bahrain'
      AND CONVERT(date, rv.created_at) > '{CUTOFF}'
)
SELECT
    r.id                                             AS reg_id,
    r.title,
    r.source_page_url,
    rv.version_id,
    rv.status                                        AS version_status,
    CASE
        WHEN CONVERT(date, rv.created_at) > '{CUTOFF}' THEN 'MONITORING'
        ELSE                                               'INITIAL CRAWL'
    END                                              AS version_origin,
    rv.change_summary,
    rv.updated_date,
    CONVERT(varchar(30), rv.created_at, 120)         AS version_created_at,
    rv.content_hash,
    CAST(rv.content_html  AS NVARCHAR(MAX))          AS content_html,
    CAST(rv.content_text  AS NVARCHAR(MAX))          AS content_text
FROM regulation_versions rv
JOIN regulations r ON r.id = rv.regulation_id
JOIN monitoring_regs m ON m.id = rv.regulation_id
ORDER BY r.id ASC, rv.version_id ASC;
"""


# ── Excel helpers ─────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(color="FFFFFF", bold=True)
NEW_FILL     = PatternFill("solid", fgColor="C6EFCE")   # green  — new reg
UPD_FILL     = PatternFill("solid", fgColor="FFEB9C")   # yellow — updated reg
MON_V_FILL   = PatternFill("solid", fgColor="BDD7EE")   # blue   — monitoring version
INIT_V_FILL  = PatternFill("solid", fgColor="F2F2F2")   # grey   — initial crawl version

def _write_sheet(ws, headers, rows, row_color_fn=None):
    # header row
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = Alignment(wrap_text=False)

    # data rows
    for r_idx, row in enumerate(rows, 2):
        fill = row_color_fn(row, headers) if row_color_fn else None
        for c_idx, val in enumerate(row, 1):
            txt = str(val) if val is not None else ""
            if len(txt) > EXCEL_MAX:
                txt = txt[:EXCEL_MAX] + "… [TRUNCATED]"
            cell = ws.cell(r_idx, c_idx, txt)
            cell.alignment = Alignment(wrap_text=False, vertical="top")
            if fill:
                cell.fill = fill

    # column widths — cap at 80, minimum 12
    for c_idx, h in enumerate(headers, 1):
        col_vals = [str(row[c_idx - 1] or "")[:120] for row in rows]
        width = max(len(h), max((len(v) for v in col_vals), default=0))
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(width + 2, 12), 80)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def s1_color(row, headers):
    action = row[headers.index("monitoring_action")]
    return NEW_FILL if "NEW" in str(action) else UPD_FILL

def s2_color(row, headers):
    origin = row[headers.index("version_origin")]
    return MON_V_FILL if "MONITORING" in str(origin) else INIT_V_FILL


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("Connecting …")
    c = conn()
    cur = c.cursor()

    print("Querying Sheet 1 (monitoring fetched) …")
    cur.execute(SHEET1_SQL)
    s1_rows = cur.fetchall()
    s1_cols = [d[0] for d in cur.description]
    print(f"  {len(s1_rows)} regulation(s) found")

    print("Querying Sheet 2 (version history) …")
    cur.execute(SHEET2_SQL)
    s2_rows = cur.fetchall()
    s2_cols = [d[0] for d in cur.description]
    print(f"  {len(s2_rows)} version row(s) found")

    c.close()

    print(f"Writing {OUTPUT} …")
    wb = openpyxl.Workbook()

    # Sheet 1
    ws1 = wb.active
    ws1.title = "Monitoring Fetched"
    _write_sheet(ws1, s1_cols, s1_rows, s1_color)

    # Sheet 2
    ws2 = wb.create_sheet("Version History")
    _write_sheet(ws2, s2_cols, s2_rows, s2_color)

    wb.save(OUTPUT)
    print(f"\nDone.")
    print(f"  Sheet 1 — Monitoring Fetched : {len(s1_rows)} rows")
    print(f"  Sheet 2 — Version History    : {len(s2_rows)} rows")
    print(f"  File: {OUTPUT}")
    print()
    print("Colour key:")
    print("  Sheet 1 — green  = NEW regulation added by monitoring")
    print("            yellow = EXISTING regulation that got a new version")
    print("  Sheet 2 — blue   = version created by monitoring")
    print("            grey   = version from initial crawl (baseline)")


if __name__ == "__main__":
    main()
