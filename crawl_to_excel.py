"""
crawl_to_excel.py
=================
Crawls CBB Rulebook volumes and saves:
  - One Excel file per volume  →  output/excel/<volume_name>.xlsx
  - One JSON cache per volume  →  output/cache/<volume_name>.json

Usage:
    python crawl_to_excel.py              # all volumes
    python crawl_to_excel.py --vol 7      # only Vol 7
"""

import json
import logging
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import List, Dict

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path("output")
EXCEL_DIR  = OUTPUT_DIR / "excel"
CACHE_DIR  = OUTPUT_DIR / "cache"
EXCEL_DIR.mkdir(parents=True, exist_ok=True)
CACHE_DIR.mkdir(parents=True, exist_ok=True)

MAX_WORKERS   = 4
REQUEST_DELAY = 1.0

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(threadName)s] %(message)s",
    datefmt="%H:%M:%S",
    stream=sys.stdout,
)
log = logging.getLogger(__name__)

try:
    from cbb_test_crawlers.cbb_rulebook_crawler import _collect_volumes, _process, RulebookDoc
except ImportError:
    from cbb_rulebook_crawler import _collect_volumes, _process, RulebookDoc


def _safe_name(text: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "_", text).strip()


def _clean_html(raw: str) -> str:
    if not raw:
        return raw
    if "\\r\\n" not in raw and "\\n" not in raw:
        return raw
    cleaned = raw
    cleaned = cleaned.replace("\\r\\n", "\n").replace("\\r", "\n").replace("\\n", "\n")
    cleaned = cleaned.replace("\\t", "\t").replace("\\u0026", "&")
    cleaned = cleaned.replace('\\"', '"').replace("\\'", "'")
    cleaned = cleaned.replace("&amp;", "&")
    return cleaned


def _html_to_text(html: str) -> str:
    if not html:
        return ""
    return BeautifulSoup(_clean_html(html), "html.parser").get_text(separator=" ", strip=True)


def crawl_volume(volume_node) -> Dict:
    name = volume_node.text.strip()
    t0   = time.time()
    results: List = []
    visited = set()
    log.info(f"START: {name}")
    try:
        _process(
            node=volume_node, path=["CBB Rulebook"], depth=0,
            visited=visited, results=results, request_delay=REQUEST_DELAY,
        )
    except Exception as e:
        log.error(f"Error crawling '{name}': {e}", exc_info=True)
    elapsed = time.time() - t0
    folders = [d for d in results if d.is_folder]
    leaves  = [d for d in results if not d.is_folder]
    log.info(f"DONE: {name} — {len(results)} docs ({len(folders)} folders, {len(leaves)} leaves) in {elapsed:.0f}s")
    return {"volume": name, "results": results, "elapsed": elapsed}


HEADERS = [
    "row_type", "depth", "indent_title", "title", "url",
    "pdf_link", "pdf_links_all", "faq_link", "category", "content_hash",
    "html_ok", "has_html", "has_text",
    "path_level_1", "path_level_2", "path_level_3", "path_level_4", "path_level_5",
    "full_path", "content_text_preview",
]

HDR_FILL   = PatternFill("solid", start_color="1F3864")
F_FILL     = PatternFill("solid", start_color="D9E1F2")
R_FILL     = PatternFill("solid", start_color="E2EFDA")
ALT_F_FILL = PatternFill("solid", start_color="C5D0E8")
ALT_R_FILL = PatternFill("solid", start_color="D0E8C5")
WARN_FILL  = PatternFill("solid", start_color="FFE4E1")
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT  = Font(name="Arial", size=9)
BOLD_FONT  = Font(name="Arial", bold=True, size=9)
THIN       = Side(border_style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _doc_to_row(doc) -> list:
    path     = doc.doc_path or []
    depth    = max(0, len(path) - 1)
    indent   = ("    " * depth) + (doc.title or "")
    row_type = "F" if doc.is_folder else "R"
    levels   = [(path[i] if i < len(path) else "") for i in range(5)]
    extra       = getattr(doc, "extra_meta", {}) or {}
    pdf_link    = extra.get("pdf_link") or ""
    all_pdfs    = extra.get("pdf_links", [])
    faq_link    = extra.get("faq_link") or ""
    all_pdfs_str = ", ".join(p["url"] if isinstance(p, dict) else p for p in all_pdfs) if all_pdfs else pdf_link
    raw_html   = getattr(doc, "document_html", "") or ""
    html_ok    = "WARN(\\r\\n)" if ("\\r\\n" in raw_html or "\\n" in raw_html) else "OK"
    clean_html = _clean_html(raw_html)
    raw_text   = getattr(doc, "content_text", "") or ""
    if "\\r\\n" in raw_text or "\\n" in raw_text:
        preview_text = _html_to_text(clean_html)
    else:
        preview_text = raw_text or _html_to_text(clean_html)
    preview = preview_text[:300].replace("\n", " ").strip()
    return [
        row_type, depth, indent, doc.title or "", doc.url or "",
        pdf_link, all_pdfs_str, faq_link,
        getattr(doc, "category", "") or "",
        getattr(doc, "content_hash", "") or "",
        html_ok,
        "YES" if raw_html else "NO",
        "YES" if raw_text else "NO",
        *levels,
        " > ".join(str(p) for p in path),
        preview,
    ]


def save_to_excel(volume_name: str, docs: list) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Crawl Results"
    folders   = sum(1 for d in docs if d.is_folder)
    leaves    = sum(1 for d in docs if not d.is_folder)
    has_pdf   = sum(1 for d in docs if (getattr(d, "extra_meta", {}) or {}).get("pdf_link"))
    warn_html = sum(1 for d in docs if not d.is_folder and "\\r\\n" in (getattr(d, "document_html", "") or ""))

    ws.append([
        f"Volume: {volume_name}", f"Total: {len(docs)}",
        f"Folders (F): {folders}", f"Regulations (R): {leaves}",
        f"With PDF link: {has_pdf}", f"HTML escape warnings: {warn_html}",
        f"Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}",
    ])
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, size=10, color="1F3864")
    ws.row_dimensions[1].height = 18
    ws.append([])

    ws.append(HEADERS)
    hdr_row = ws.max_row
    for cell in ws[hdr_row]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[hdr_row].height = 30

    for i, doc in enumerate(docs):
        row_data = _doc_to_row(doc)
        ws.append(row_data)
        row_idx  = ws.max_row
        is_folder = doc.is_folder
        alt = (i % 2 == 1)
        has_warn = row_data[10] != "OK"
        if is_folder:
            fill = ALT_F_FILL if alt else F_FILL
            font = BOLD_FONT
        elif has_warn:
            fill = WARN_FILL
            font = BODY_FONT
        else:
            fill = ALT_R_FILL if alt else R_FILL
            font = BODY_FONT
        for cell in ws[row_idx]:
            cell.fill = fill
            cell.font = font
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        ws.row_dimensions[row_idx].height = 15

    col_widths = {1:10, 2:6, 3:52, 4:40, 5:55, 6:55, 7:70, 8:45,
                  9:20, 10:35, 11:16, 12:10, 13:10, 14:25, 15:25,
                  16:22, 17:18, 18:18, 19:70, 20:60}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = f"A{hdr_row + 1}"
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(HEADERS))}{hdr_row}"

    out_path = EXCEL_DIR / f"{_safe_name(volume_name)}.xlsx"
    wb.save(str(out_path))
    log.info(f"Saved Excel → {out_path}")
    return out_path


def save_to_cache(volume_name: str, docs: list) -> Path:
    records = []
    for doc in docs:
        raw_html   = getattr(doc, "document_html", "") or ""
        clean_html = _clean_html(raw_html)
        raw_text   = getattr(doc, "content_text", "") or ""
        if "\\r\\n" in raw_text or "\\n" in raw_text:
            clean_text = _html_to_text(clean_html)
        else:
            clean_text = raw_text or _html_to_text(clean_html)
        extra = getattr(doc, "extra_meta", {}) or {}
        records.append({
            "is_folder":     doc.is_folder,
            "title":         doc.title or "",
            "url":           doc.url or "",
            "doc_path":      doc.doc_path,
            "document_html": clean_html,
            "content_text":  clean_text,
            "content_hash":  getattr(doc, "content_hash", "") or "",
            "pdf_link":      extra.get("pdf_link") or "",
            "pdf_links":     extra.get("pdf_links") or [],
            "faq_link":      extra.get("faq_link") or "",
            "extra_meta":    extra,
        })
    out_path = CACHE_DIR / f"{_safe_name(volume_name)}.json"
    out_path.write_text(json.dumps(records, indent=2, ensure_ascii=False))
    log.info(f"Saved cache → {out_path}")
    return out_path


def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--vol", type=str, default=None, help="Filter by volume number e.g. 7")
    args = parser.parse_args()

    t_start = time.time()
    log.info("=== CBB RULEBOOK — CRAWL TO EXCEL ===\n")
    volumes = _collect_volumes("https://cbben.thomsonreuters.com/rulebook/common-volume")
    if not volumes:
        log.error("No volumes found — exiting.")
        return
    if args.vol:
        volumes = [v for v in volumes if args.vol in v.text or f"Volume {args.vol}" in v.text or "Collective" in v.text and args.vol == "7"]
        log.info(f"Filtered to {len(volumes)} volume(s) matching --vol {args.vol}")

    log.info(f"Crawling {len(volumes)} volume(s)\n")
    outcomes = []
    with ThreadPoolExecutor(max_workers=min(MAX_WORKERS, len(volumes)), thread_name_prefix="crawler") as pool:
        futures = {pool.submit(crawl_volume, v): v for v in volumes}
        for future in as_completed(futures):
            try:
                outcome = future.result()
                outcomes.append(outcome)
                save_to_excel(outcome["volume"], outcome["results"])
                save_to_cache(outcome["volume"], outcome["results"])
            except Exception as e:
                log.error(f"Volume failed: {e}", exc_info=True)

    log.info(f"\nTotal time: {time.time()-t_start:.0f}s")
    log.info(f"Excel → {EXCEL_DIR.resolve()}")
    log.info(f"Cache → {CACHE_DIR.resolve()}")


if __name__ == "__main__":
    main()