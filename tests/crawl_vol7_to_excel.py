"""
crawl_vol7_to_excel.py
======================
Crawls ONLY "Volume 7 — Collective Investment Undertakings" from the CBB
Rulebook sidebar and saves:

  - Excel (with F/R leaf-parent hierarchy)  →  output/excel/vol7_pending.xlsx
  - JSON cache (for the DB-push step)        →  output/cache/vol7_pending.json

Run this first.  Review the Excel file.  When satisfied run:
    python tests/push_vol7_to_db.py

Usage:
    python tests/crawl_vol7_to_excel.py
"""

import json
import logging
import re
import sys
import time
from pathlib import Path
from typing import List, Dict

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ──────────────────────────────────────────────────────────────────────
ROOT       = Path(__file__).resolve().parent.parent
OUTPUT_DIR = ROOT / "output"
EXCEL_DIR  = OUTPUT_DIR / "excel"
CACHE_DIR  = OUTPUT_DIR / "cache"
EXCEL_DIR.mkdir(parents=True, exist_ok=True)
CACHE_DIR.mkdir(parents=True, exist_ok=True)

EXCEL_OUT = EXCEL_DIR / "vol7_pending.xlsx"
CACHE_OUT = CACHE_DIR / "vol7_pending.json"

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(message)s",
    datefmt="%H:%M:%S",
    stream=sys.stdout,
)
log = logging.getLogger(__name__)

# ── Crawler imports ────────────────────────────────────────────────────────────
sys.path.insert(0, str(ROOT))
from cbb_test_crawlers.cbb_rulebook_crawler import (
    _collect_volumes,
    _process,
    RulebookDoc,
    SIDEBAR_SEED,
)

# ── Volume 7 keyword ───────────────────────────────────────────────────────────
VOL7_KEYWORD = "volume 7"   # case-insensitive match


# ── HTML helpers ───────────────────────────────────────────────────────────────

def _clean_html(raw: str) -> str:
    """Decode literal \\r\\n escape sequences left by JSON pass-through."""
    if not raw or ("\\r\\n" not in raw and "\\n" not in raw):
        return raw
    cleaned = raw
    for src, dst in [
        ("\\r\\n", "\n"), ("\\r",    "\n"), ("\\n",    "\n"),
        ("\\t",    "\t"), ("\\u0026", "&"), ("\\\"",    '"'),
        ("\\'",     "'"), ("&amp;",   "&"),
    ]:
        cleaned = cleaned.replace(src, dst)
    return cleaned


def _html_to_text(html: str) -> str:
    if not html:
        return ""
    return BeautifulSoup(_clean_html(html), "html.parser").get_text(separator=" ", strip=True)


# ── Excel styles ───────────────────────────────────────────────────────────────

HEADERS = [
    "row_type",          # F = Folder  |  R = Regulation/Leaf
    "depth",             # 0 = volume root
    "indent_title",      # title visually indented for hierarchy
    "title",
    "url",
    "pdf_link",
    "pdf_links_all",
    "faq_link",
    "content_hash",
    "html_ok",           # OK  or  WARN(\r\n)
    "has_html",
    "has_text",
    "path_level_1",
    "path_level_2",
    "path_level_3",
    "path_level_4",
    "path_level_5",
    "full_path",
    "content_text_preview",
]

HDR_FILL    = PatternFill("solid", start_color="1F3864")
F_FILL      = PatternFill("solid", start_color="D9E1F2")   # blue  → folders
R_FILL      = PatternFill("solid", start_color="E2EFDA")   # green → leaves
ALT_F_FILL  = PatternFill("solid", start_color="C5D0E8")
ALT_R_FILL  = PatternFill("solid", start_color="D0E8C5")
WARN_FILL   = PatternFill("solid", start_color="FFE4E1")   # pink  → HTML issues
HDR_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT   = Font(name="Arial", size=9)
BOLD_FONT   = Font(name="Arial", bold=True, size=9)
THIN        = Side(border_style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COL_WIDTHS  = {
    1: 10, 2: 6, 3: 52, 4: 40, 5: 55, 6: 55, 7: 70, 8: 45,
    9: 35, 10: 16, 11: 10, 12: 10,
    13: 25, 14: 25, 15: 22, 16: 18, 17: 18,
    18: 70, 19: 60,
}


def _doc_to_row(doc: RulebookDoc) -> list:
    path     = doc.doc_path or []
    depth    = max(0, len(path) - 1)
    indent   = ("    " * depth) + (doc.title or "")
    row_type = "F" if doc.is_folder else "R"
    levels   = [(path[i] if i < len(path) else "") for i in range(5)]

    extra       = getattr(doc, "extra_meta", {}) or {}
    pdf_link    = extra.get("pdf_link") or ""
    all_pdfs    = extra.get("pdf_links", [])
    faq_link    = extra.get("faq_link") or ""
    all_pdfs_str = (
        ", ".join(p["url"] if isinstance(p, dict) else p for p in all_pdfs)
        if all_pdfs else pdf_link
    )

    raw_html  = getattr(doc, "document_html", "") or ""
    html_ok   = "WARN(\\r\\n)" if ("\\r\\n" in raw_html or "\\n" in raw_html) else "OK"
    clean_html = _clean_html(raw_html)

    raw_text = getattr(doc, "content_text", "") or ""
    preview_src = (
        _html_to_text(clean_html)
        if ("\\r\\n" in raw_text or "\\n" in raw_text)
        else (raw_text or _html_to_text(clean_html))
    )
    preview = preview_src[:300].replace("\n", " ").strip()

    return [
        row_type, depth, indent, doc.title or "", doc.url or "",
        pdf_link, all_pdfs_str, faq_link,
        getattr(doc, "content_hash", "") or "",
        html_ok,
        "YES" if raw_html  else "NO",
        "YES" if raw_text  else "NO",
        *levels,
        " > ".join(str(p) for p in path),
        preview,
    ]


def save_excel(docs: List[RulebookDoc], vol_name: str) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vol7 Crawl"

    folders   = sum(1 for d in docs if d.is_folder)
    leaves    = sum(1 for d in docs if not d.is_folder)
    has_pdf   = sum(1 for d in docs if (getattr(d, "extra_meta", {}) or {}).get("pdf_link"))
    warn_html = sum(
        1 for d in docs
        if not d.is_folder and "\\r\\n" in (getattr(d, "document_html", "") or "")
    )

    # Summary row
    ws.append([
        f"Volume: {vol_name}",
        f"Total: {len(docs)}",
        f"Folders (F): {folders}",
        f"Regulations (R): {leaves}",
        f"With PDF: {has_pdf}",
        f"HTML warnings: {warn_html}",
        f"Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}",
    ])
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, size=10, color="1F3864")
    if warn_html:
        ws["F1"].font = Font(name="Arial", bold=True, size=10, color="FF0000")
    ws.row_dimensions[1].height = 18
    ws.append([])   # spacer

    # Header row
    ws.append(HEADERS)
    hdr_row = ws.max_row
    for cell in ws[hdr_row]:
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
    ws.row_dimensions[hdr_row].height = 30

    # Data rows
    for i, doc in enumerate(docs):
        row_data = _doc_to_row(doc)
        ws.append(row_data)
        row_idx  = ws.max_row
        alt      = (i % 2 == 1)
        has_warn = row_data[9] != "OK"   # html_ok index in HEADERS

        if doc.is_folder:
            fill, font = (ALT_F_FILL if alt else F_FILL), BOLD_FONT
        elif has_warn:
            fill, font = WARN_FILL, BODY_FONT
        else:
            fill, font = (ALT_R_FILL if alt else R_FILL), BODY_FONT

        for cell in ws[row_idx]:
            cell.fill      = fill
            cell.font      = font
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        ws.row_dimensions[row_idx].height = 15

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = f"A{hdr_row + 1}"
    ws.auto_filter.ref = (
        f"A{hdr_row}:{get_column_letter(len(HEADERS))}{hdr_row}"
    )

    # Legend sheet
    lg = wb.create_sheet("Legend")
    for row in [
        ["Column",               "Meaning"],
        ["row_type",             "F = Folder (stored in compliancecategory)  |  R = Regulation (stored in regulations)"],
        ["depth",                "0 = volume root, 1 = first-level child, etc."],
        ["indent_title",         "Title indented to show folder tree hierarchy"],
        ["pdf_link",             "Primary PDF URL extracted from the page"],
        ["pdf_links_all",        "All PDF URLs found (comma-separated)"],
        ["faq_link",             "FAQ URL if found"],
        ["html_ok",              "OK = clean HTML.  WARN(\\r\\n) = had JSON escape sequences (auto-fixed in cache)"],
        ["has_html / has_text",  "Whether document_html / content_text was captured"],
        ["path_level_1..5",      "Individual path segments for each hierarchy level"],
        ["full_path",            "All segments joined with ' > '"],
        ["content_text_preview", "First 300 chars of clean plain text"],
        [],
        ["Colours"],
        ["Blue (bold)",  "Folders  →  become compliancecategory entries in the DB"],
        ["Green",        "Regulations / leaf articles  →  go into the regulations table"],
        ["Pink",         "Leaves with HTML escape issues (auto-fixed before DB push)"],
        [],
        ["After reviewing this file run:"],
        ["  python tests/push_vol7_to_db.py"],
    ]:
        lg.append(row)
    lg.column_dimensions["A"].width = 28
    lg.column_dimensions["B"].width = 80

    wb.save(str(EXCEL_OUT))
    log.info(f"Saved Excel  →  {EXCEL_OUT}")
    return EXCEL_OUT


def save_cache(docs: List[RulebookDoc], vol_name: str) -> Path:
    records = []
    for doc in docs:
        raw_html   = getattr(doc, "document_html", "") or ""
        clean_html = _clean_html(raw_html)

        raw_text   = getattr(doc, "content_text", "") or ""
        clean_text = (
            _html_to_text(clean_html)
            if ("\\r\\n" in raw_text or "\\n" in raw_text)
            else (raw_text or _html_to_text(clean_html))
        )

        extra = getattr(doc, "extra_meta", {}) or {}

        records.append({
            "is_folder":     doc.is_folder,
            "row_type":      "F" if doc.is_folder else "R",
            "title":         doc.title or "",
            "url":           doc.url or "",
            "doc_path":      doc.doc_path,
            "depth":         max(0, len(doc.doc_path) - 1),
            "document_html": clean_html,
            "content_text":  clean_text,
            "content_hash":  getattr(doc, "content_hash", "") or "",
            "pdf_link":      extra.get("pdf_link")  or "",
            "pdf_links":     extra.get("pdf_links") or [],
            "faq_link":      extra.get("faq_link")  or "",
            "extra_meta":    extra,
        })

    CACHE_OUT.write_text(
        json.dumps({"volume": vol_name, "crawled_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
                    "docs": records},
                   indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    log.info(f"Saved cache  →  {CACHE_OUT}")
    return CACHE_OUT


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    t0 = time.time()
    log.info("=" * 60)
    log.info("CBB RULEBOOK — VOLUME 7 CRAWL")
    log.info("=" * 60)

    # ── 1. Discover all volumes ────────────────────────────────────────────────
    log.info("Fetching volume list from sidebar …")
    volumes = _collect_volumes(SIDEBAR_SEED)
    if not volumes:
        log.error("No volumes found. Check network/URL.")
        sys.exit(1)

    log.info(f"Found {len(volumes)} total volumes:")
    for i, v in enumerate(volumes, 1):
        log.info(f"  {i:2d}. {v.text}")

    # ── 2. Filter to Volume 7 ─────────────────────────────────────────────────
    vol7_nodes = [v for v in volumes if VOL7_KEYWORD in v.text.lower()]
    if not vol7_nodes:
        log.error(f"Could not find a volume matching '{VOL7_KEYWORD}'. Volumes found:")
        for v in volumes:
            log.error(f"  {v.text!r}")
        sys.exit(1)

    vol7 = vol7_nodes[0]
    log.info(f"\nTarget volume: {vol7.text!r}\n")

    # ── 3. Crawl ──────────────────────────────────────────────────────────────
    results: List[RulebookDoc] = []
    visited = set()
    log.info("Starting crawl (this may take several minutes) …")
    _process(
        node=vol7,
        path=["CBB Rulebook"],
        depth=0,
        visited=visited,
        results=results,
        request_delay=1.2,
    )

    folders = [d for d in results if d.is_folder]
    leaves  = [d for d in results if not d.is_folder]
    elapsed = time.time() - t0

    log.info(f"\nCrawl complete in {elapsed:.0f}s")
    log.info(f"  Total docs : {len(results)}")
    log.info(f"  Folders    : {len(folders)}")
    log.info(f"  Leaves     : {len(leaves)}")

    if not results:
        log.error("No documents crawled — check connectivity.")
        sys.exit(1)

    # ── 4. Save Excel ─────────────────────────────────────────────────────────
    save_excel(results, vol7.text)

    # ── 5. Save JSON cache ────────────────────────────────────────────────────
    save_cache(results, vol7.text)

    log.info("\n" + "=" * 60)
    log.info("DONE")
    log.info(f"  Excel  →  {EXCEL_OUT}")
    log.info(f"  Cache  →  {CACHE_OUT}")
    log.info("")
    log.info("Review the Excel file.  When satisfied run:")
    log.info("  python tests/push_vol7_to_db.py")
    log.info("=" * 60)


if __name__ == "__main__":
    main()
