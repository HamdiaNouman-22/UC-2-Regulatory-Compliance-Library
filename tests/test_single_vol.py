"""
crawl_to_excel.py  (FIXED)
==========================
Crawls CBB Rulebook volumes in parallel and saves:
  - One Excel file per volume  →  output/excel/<volume_name>.xlsx
  - One JSON cache per volume  →  output/cache/<volume_name>.json

Fixes vs original:
  1. HTML stored as literal \\r\\n strings → cleaned before cache/Excel
  2. PDF links (extra_meta) now saved in cache and shown in Excel
  3. content_text re-derived from clean HTML if it had escape issues
  4. Pink row highlight in Excel for any rows with HTML escape warnings

Run this first. Review the Excel files. When satisfied, run push_to_db.py.

Usage:
    python crawl_to_excel.py
"""

import json
import logging
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import List, Dict

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── PATHS ──────────────────────────────────────────────────────────────────────
OUTPUT_DIR = Path("output")
EXCEL_DIR  = OUTPUT_DIR / "excel"
CACHE_DIR  = OUTPUT_DIR / "cache"
EXCEL_DIR.mkdir(parents=True, exist_ok=True)
CACHE_DIR.mkdir(parents=True, exist_ok=True)

# ── CONFIG ─────────────────────────────────────────────────────────────────────
MAX_WORKERS   = 4
REQUEST_DELAY = 1.0

# ── LOGGING ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(threadName)s] %(message)s",
    datefmt="%H:%M:%S",
    stream=sys.stdout,
)
log = logging.getLogger(__name__)

from cbb_test_crawlers.cbb_rulebook_crawler import (
    _collect_volumes,
    _process,
    RulebookDoc,
)


# ── SAFE FILENAME ──────────────────────────────────────────────────────────────
def _safe_name(text: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "_", text).strip()


# ── HTML CLEANING ──────────────────────────────────────────────────────────────
def _clean_html(raw: str) -> str:
    """
    Fix HTML that contains literal \\r\\n escape sequences instead of real
    newlines. This happens when content was fetched from a JSON API and the
    string value was stored directly without decoding escape sequences.

    Also handles: \\n, \\t, \\u0026, doubled &amp; from JSON pass-through.
    Safe to call on already-clean HTML — returns unchanged if no escapes found.
    """
    if not raw:
        return raw
    if "\\r\\n" not in raw and "\\n" not in raw:
        return raw  # already clean — skip processing entirely
    cleaned = raw
    cleaned = cleaned.replace("\\r\\n", "\n")
    cleaned = cleaned.replace("\\r",    "\n")
    cleaned = cleaned.replace("\\n",    "\n")
    cleaned = cleaned.replace("\\t",    "\t")
    cleaned = cleaned.replace("\\u0026", "&")
    cleaned = cleaned.replace("\\\"",    '"')
    cleaned = cleaned.replace("\\'",     "'")
    cleaned = cleaned.replace("&amp;",   "&")
    return cleaned


def _html_to_text(html: str) -> str:
    """Extract plain text from HTML string (after cleaning escape sequences)."""
    if not html:
        return ""
    return BeautifulSoup(_clean_html(html), "html.parser").get_text(separator=" ", strip=True)


# ── CRAWL ONE VOLUME ───────────────────────────────────────────────────────────
def crawl_volume(volume_node) -> Dict:
    name = volume_node.text.strip()
    t0   = time.time()

    results: List[RulebookDoc] = []
    visited = set()

    log.info(f"START: {name}")
    try:
        _process(
            node=volume_node,
            path=["CBB Rulebook"],
            depth=0,
            visited=visited,
            results=results,
            request_delay=REQUEST_DELAY,
        )
    except Exception as e:
        log.error(f"Error crawling '{name}': {e}", exc_info=True)

    elapsed = time.time() - t0
    folders = [d for d in results if d.is_folder]
    leaves  = [d for d in results if not d.is_folder]
    log.info(f"DONE: {name} — {len(results)} docs ({len(folders)} folders, {len(leaves)} leaves) in {elapsed:.0f}s")

    return {"volume": name, "results": results, "elapsed": elapsed}


# ── EXCEL STYLES ───────────────────────────────────────────────────────────────
HEADERS = [
    "row_type",               # F = Folder, R = Regulation/Leaf
    "depth",                  # nesting depth (0 = root)
    "indent_title",           # title indented visually for hierarchy
    "title",
    "url",
    "pdf_link",               # ← primary PDF link
    "pdf_links_all",          # ← all PDFs comma-separated
    "faq_link",               # ← FAQ link if present
    "category",
    "content_hash",
    "html_ok",                # OK or WARN(\r\n)
    "has_html",               # YES / NO
    "has_text",               # YES / NO
    "path_level_1",
    "path_level_2",
    "path_level_3",
    "path_level_4",
    "path_level_5",
    "full_path",              # full path joined with " > "
    "content_text_preview",   # first 300 chars of clean text
]

HDR_FILL    = PatternFill("solid", start_color="1F3864")
F_FILL      = PatternFill("solid", start_color="D9E1F2")
R_FILL      = PatternFill("solid", start_color="E2EFDA")
ALT_F_FILL  = PatternFill("solid", start_color="C5D0E8")
ALT_R_FILL  = PatternFill("solid", start_color="D0E8C5")
WARN_FILL   = PatternFill("solid", start_color="FFE4E1")   # pink for HTML warn rows
HDR_FONT    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT   = Font(name="Arial", size=9)
BOLD_FONT   = Font(name="Arial", bold=True, size=9)
THIN        = Side(border_style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _doc_to_row(doc: RulebookDoc) -> list:
    path      = doc.doc_path or []
    depth     = max(0, len(path) - 1)
    indent    = ("    " * depth) + (doc.title or "")
    row_type  = "F" if doc.is_folder else "R"
    levels    = [(path[i] if i < len(path) else "") for i in range(5)]

    # ── PDF / extra_meta ──────────────────────────────────────────────────────
    # RulebookDoc from cbb_rulebook_crawler stores pdf_link in extra_meta
    extra       = getattr(doc, "extra_meta", {}) or {}
    pdf_link    = extra.get("pdf_link") or ""
    all_pdfs    = extra.get("pdf_links", [])
    faq_link    = extra.get("faq_link") or ""
    all_pdfs_str = (
        ", ".join(p["url"] if isinstance(p, dict) else p for p in all_pdfs)
        if all_pdfs else pdf_link
    )

    # ── HTML quality check ────────────────────────────────────────────────────
    raw_html  = getattr(doc, "document_html", "") or ""
    html_ok   = "WARN(\\r\\n)" if ("\\r\\n" in raw_html or "\\n" in raw_html) else "OK"
    clean_html = _clean_html(raw_html)

    # ── Clean text preview ────────────────────────────────────────────────────
    raw_text = getattr(doc, "content_text", "") or ""
    if "\\r\\n" in raw_text or "\\n" in raw_text:
        # text also has escape issues — re-derive from clean HTML
        preview_text = _html_to_text(clean_html)
    else:
        preview_text = raw_text or _html_to_text(clean_html)
    preview = preview_text[:300].replace("\n", " ").strip()

    return [
        row_type,
        depth,
        indent,
        doc.title or "",
        doc.url or "",
        pdf_link,
        all_pdfs_str,
        faq_link,
        getattr(doc, "category", "") or "",
        getattr(doc, "content_hash", "") or "",
        html_ok,
        "YES" if raw_html   else "NO",
        "YES" if raw_text   else "NO",
        *levels,
        " > ".join(str(p) for p in path),
        preview,
    ]


def save_to_excel(volume_name: str, docs: List[RulebookDoc]) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Crawl Results"

    folders   = sum(1 for d in docs if d.is_folder)
    leaves    = sum(1 for d in docs if not d.is_folder)
    has_pdf   = sum(1 for d in docs if (getattr(d, "extra_meta", {}) or {}).get("pdf_link"))
    warn_html = sum(
        1 for d in docs
        if not d.is_folder and (
            "\\r\\n" in (getattr(d, "document_html", "") or "")
            or "\\n"  in (getattr(d, "document_html", "") or "")
        )
    )

    # ── Summary row ────────────────────────────────────────────────────────────
    ws.append([
        f"Volume: {volume_name}",
        f"Total: {len(docs)}",
        f"Folders (F): {folders}",
        f"Regulations (R): {leaves}",
        f"With PDF link: {has_pdf}",
        f"HTML escape warnings: {warn_html}",
        f"Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}",
    ])
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, size=10, color="1F3864")
    if warn_html > 0:
        ws["F1"].font = Font(name="Arial", bold=True, size=10, color="FF0000")
    ws.row_dimensions[1].height = 18
    ws.append([])   # blank spacer

    # ── Headers ────────────────────────────────────────────────────────────────
    ws.append(HEADERS)
    hdr_row = ws.max_row
    for cell in ws[hdr_row]:
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
    ws.row_dimensions[hdr_row].height = 30

    # ── Data rows ──────────────────────────────────────────────────────────────
    for i, doc in enumerate(docs):
        row_data  = _doc_to_row(doc)
        ws.append(row_data)
        row_idx   = ws.max_row
        is_folder = doc.is_folder
        alt       = (i % 2 == 1)
        has_warn  = row_data[10] != "OK"   # html_ok is col index 10 (0-based)

        if is_folder:
            fill = ALT_F_FILL if alt else F_FILL
            font = BOLD_FONT
        elif has_warn:
            fill = WARN_FILL    # pink — HTML had escape sequences
            font = BODY_FONT
        else:
            fill = ALT_R_FILL if alt else R_FILL
            font = BODY_FONT

        for cell in ws[row_idx]:
            cell.fill      = fill
            cell.font      = font
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

        ws.row_dimensions[row_idx].height = 15

    # ── Column widths ──────────────────────────────────────────────────────────
    col_widths = {
        1:  10,   # row_type
        2:  6,    # depth
        3:  52,   # indent_title
        4:  40,   # title
        5:  55,   # url
        6:  55,   # pdf_link
        7:  70,   # pdf_links_all
        8:  45,   # faq_link
        9:  20,   # category
        10: 35,   # content_hash
        11: 16,   # html_ok
        12: 10,   # has_html
        13: 10,   # has_text
        14: 25,   # path levels
        15: 25,
        16: 22,
        17: 18,
        18: 18,
        19: 70,   # full_path
        20: 60,   # preview
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze pane + autofilter ────────────────────────────────────────────────
    ws.freeze_panes = f"A{hdr_row + 1}"
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(HEADERS))}{hdr_row}"

    # ── Legend sheet ───────────────────────────────────────────────────────────
    lg = wb.create_sheet("Legend")
    for r in [
        ["Column",            "Meaning"],
        ["row_type",          "F = Folder (ComplianceCategory, type=F), R = Regulation (type=R)"],
        ["depth",             "Nesting depth. 0 = root, 1 = first child, etc."],
        ["indent_title",      "Title indented to visualise the folder tree hierarchy"],
        ["pdf_link",          "Primary PDF download URL extracted from the page"],
        ["pdf_links_all",     "All PDF URLs found (comma-separated) if multiple"],
        ["faq_link",          "FAQ link URL if found on the page"],
        ["html_ok",           "OK = clean HTML.  WARN(\\r\\n) = had JSON escape sequences (auto-fixed in cache)"],
        ["has_html",          "Whether document_html content was captured"],
        ["has_text",          "Whether content_text was extracted"],
        ["path_level_1..5",   "Individual path segments for each hierarchy level"],
        ["full_path",         "All path segments joined with ' > '"],
        ["content_text_preview", "First 300 characters of clean plain text"],
        [],
        ["Row colours"],
        ["Blue rows",  "Folders → stored as ComplianceCategory (type='F')"],
        ["Green rows", "Regulations → stored as RegulatoryDocument (type='R')"],
        ["Pink rows",  "Regulations with HTML escape issues — fixed automatically in the JSON cache"],
    ]:
        lg.append(r)
    lg.column_dimensions["A"].width = 25
    lg.column_dimensions["B"].width = 78

    out_path = EXCEL_DIR / f"{_safe_name(volume_name)}.xlsx"
    wb.save(str(out_path))
    log.info(f"Saved Excel → {out_path}  (pdf={has_pdf}, warn_html={warn_html})")
    return out_path


# ── JSON CACHE WRITER ──────────────────────────────────────────────────────────
def save_to_cache(volume_name: str, docs: List[RulebookDoc]) -> Path:
    records = []
    for doc in docs:
        raw_html   = getattr(doc, "document_html", "") or ""
        clean_html = _clean_html(raw_html)      # ← fix \r\n BEFORE writing to cache

        raw_text   = getattr(doc, "content_text", "") or ""
        if "\\r\\n" in raw_text or "\\n" in raw_text:
            clean_text = _html_to_text(clean_html)
        else:
            clean_text = raw_text or _html_to_text(clean_html)

        # ── extra_meta — includes pdf_link, pdf_links, faq_link ──────────────
        extra = getattr(doc, "extra_meta", {}) or {}

        records.append({
            "is_folder":     doc.is_folder,
            "row_type":      "F" if doc.is_folder else "R",
            "title":         doc.title or "",
            "url":           doc.url or "",
            "doc_path":      doc.doc_path,
            "category":      getattr(doc, "category", "") or "",
            "document_html": clean_html,         # ← CLEAN HTML stored in cache
            "content_text":  clean_text,         # ← clean text
            "content_hash":  getattr(doc, "content_hash", "") or "",
            # ── PDF / link meta ───────────────────────────────────────────────
            "pdf_link":      extra.get("pdf_link")  or "",
            "pdf_links":     extra.get("pdf_links") or [],
            "faq_link":      extra.get("faq_link")  or "",
            "extra_meta":    extra,              # ← full extra_meta preserved
        })

    out_path = CACHE_DIR / f"{_safe_name(volume_name)}.json"
    out_path.write_text(json.dumps(records, indent=2, ensure_ascii=False))
    log.info(f"Saved cache → {out_path}")
    return out_path


# ── MAIN ───────────────────────────────────────────────────────────────────────
def main():
    t_start = time.time()
    log.info("=== CBB RULEBOOK — CRAWL TO EXCEL (FIXED) ===\n")

    volumes = _collect_volumes("https://cbben.thomsonreuters.com/rulebook/common-volume")
    if not volumes:
        log.error("No volumes found — exiting.")
        return

    log.info(f"Found {len(volumes)} volumes\n")

    outcomes = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS, thread_name_prefix="crawler") as pool:
        futures = {pool.submit(crawl_volume, v): v for v in volumes}
        for future in as_completed(futures):
            try:
                outcome = future.result()
                outcomes.append(outcome)

                vol_name = outcome["volume"]
                docs     = outcome["results"]

                save_to_excel(vol_name, docs)
                save_to_cache(vol_name, docs)

            except Exception as e:
                log.error(f"Volume failed: {e}", exc_info=True)

    total = time.time() - t_start
    log.info(f"\n{'='*60}")
    log.info("SUMMARY")
    log.info(f"{'='*60}")
    for o in sorted(outcomes, key=lambda x: x["volume"]):
        docs      = o["results"]
        f         = sum(1 for d in docs if d.is_folder)
        r         = sum(1 for d in docs if not d.is_folder)
        pdf       = sum(1 for d in docs if (getattr(d, "extra_meta", {}) or {}).get("pdf_link"))
        warn      = sum(
            1 for d in docs
            if not d.is_folder and "\\r\\n" in (getattr(d, "document_html", "") or "")
        )
        log.info(
            f"  {o['volume'][:35]:35s} | F={f:3d} | R={r:4d} | "
            f"pdf={pdf:3d} | html_warn={warn} | t={o['elapsed']:.0f}s"
        )
    log.info(f"{'='*60}")
    log.info(f"Total time: {total:.0f}s ({total/60:.1f} min)")
    log.info(f"\nExcel files → {EXCEL_DIR.resolve()}")
    log.info(f"JSON cache  → {CACHE_DIR.resolve()}")
    log.info("\nReview Excel files. When satisfied, run:  python push_to_db.py")


if __name__ == "__main__":
    main()