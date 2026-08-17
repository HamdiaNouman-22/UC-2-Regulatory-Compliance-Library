"""
sama_finance_sector_to_excel.py
================================
Reads the already-crawled sama_finance_sector_requests.json (output of
crawler/sama_finance_sector_crawler.py) and writes a hierarchy-aware Excel
report to output/excel/SAMA_Finance_Sector.xlsx.

Hierarchy is shown two ways, matching the convention already used by
crawl_to_excel.py for the CBB rulebook reports:
  - an indented title column for quick visual scanning
  - separate "Level N" columns (one per depth under Finance Sector) so the
    sheet can be filtered/grouped by category in Excel

document_html is not embedded directly -- some documents are ~1MB of HTML,
well past Excel's 32,767 char-per-cell limit. Instead each row gets a plain
text preview plus the document_url/PDF link; the full HTML lives in the
source JSON (and per-document .html files written under output/html/).

Usage:
    python sama_finance_sector_to_excel.py
    python sama_finance_sector_to_excel.py --input sama_finance_sector_requests.json
"""

import argparse
import json
import re
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path("output")
EXCEL_DIR = OUTPUT_DIR / "excel"
HTML_DIR = OUTPUT_DIR / "html" / "sama_finance_sector"  # default; overridden per-sector
EXCEL_DIR.mkdir(parents=True, exist_ok=True)

MAX_LEVELS = 8  # "SAMA Rulebook" + "Finance Sector" + deepest category suffix (observed depth 6) = 8
PREVIEW_CHARS = 600

HEADERS = [
    "#", "Depth", "Indented Title",
    *[f"Level {i}" for i in range(1, MAX_LEVELS + 1)],
    "Title", "Reference No", "Published Date", "Hijri Date", "Year",
    "Status", "File Type", "Has PDF", "PDF Link", "Document URL",
    "HTML Length", "HTML File", "Content Preview",
]

HDR_FILL = PatternFill("solid", start_color="1F3864")
ROW_FILL = PatternFill("solid", start_color="E2EFDA")
ALT_ROW_FILL = PatternFill("solid", start_color="D0E8C5")
NO_PDF_FILL = PatternFill("solid", start_color="FFE4E1")
HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=9)
LINK_FONT = Font(name="Arial", size=9, color="0563C1", underline="single")
THIN = Side(border_style="thin", color="CCCCCC")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _safe_filename(text: str, max_len: int = 80) -> str:
    name = re.sub(r'[\\/*?:"<>|\x00-\x1f]', "_", text).strip()
    return name[:max_len] if len(name) > max_len else name


def _html_to_text(html: str) -> str:
    if not html:
        return ""
    return BeautifulSoup(html, "html.parser").get_text(separator=" ", strip=True)


def _write_html_file(doc: dict, index: int, html_dir: Path = None) -> str:
    """Write the full document_html to its own file; return the relative path (or '' if empty)."""
    html = doc.get("document_html") or ""
    if not html:
        return ""
    out_dir = html_dir if html_dir is not None else HTML_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    fname = f"{index:04d}_{_safe_filename(doc['title'])}.html"
    path = out_dir / fname
    title = doc.get("title") or ""
    wrapped = (
        f'<!DOCTYPE html><html><head>'
        f'<meta charset="utf-8">'
        f'<base href="https://rulebook.sama.gov.sa/">'
        f'<title>{title}</title>'
        f'<style>body{{font-family:Arial,sans-serif;max-width:960px;margin:2rem auto;padding:0 1rem;line-height:1.6}}</style>'
        f'</head><body>{html}</body></html>'
    )
    path.write_text(wrapped, encoding="utf-8")
    return str(path)


def _doc_to_row(doc: dict, index: int, html_file: str, sector_label: str = "Finance Sector") -> list:
    doc_path = doc.get("doc_path") or []
    # doc_path = ["SAMA", "SAMA RULEBOOK", "Finance Sector", <category>, ..., <title>]
    # Collapse the regulator/source_system pair into one "SAMA Rulebook" level so the
    # hierarchy reads top-down: SAMA Rulebook > Finance Sector > Laws and Regulations > ...
    full_path = ["SAMA Rulebook"] + doc_path[2:] if len(doc_path) > 2 else doc_path
    depth = len(full_path)
    indented_title = ("    " * max(0, depth - 1)) + (doc.get("title") or "")
    levels = [(full_path[i] if i < len(full_path) else "") for i in range(MAX_LEVELS)]

    extra = doc.get("extra_meta") or {}
    pdf_link = extra.get("org_pdf_link") or ""
    html = doc.get("document_html") or ""
    preview = _html_to_text(html)[:PREVIEW_CHARS]

    return [
        index, depth, indented_title,
        *levels,
        doc.get("title") or "",
        doc.get("reference_no") or "",
        doc.get("published_date") or "",
        extra.get("issue_date_hijri") or "",
        doc.get("year") or "",
        extra.get("status") or "",
        doc.get("file_type") or "",
        "YES" if pdf_link else "NO",
        pdf_link,
        doc.get("document_url") or "",
        len(html),
        html_file,
        preview,
    ]


def build_excel(documents: list, out_path: Path, sector_label: str = "Finance Sector") -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = sector_label[:31]  # Excel sheet name max 31 chars

    by_category = {}
    for d in documents:
        path = d.get("doc_path") or []
        cat = path[3] if len(path) > 3 else "(uncategorized)"
        by_category[cat] = by_category.get(cat, 0) + 1
    with_pdf = sum(1 for d in documents if (d.get("extra_meta") or {}).get("org_pdf_link"))

    ws.append([
        f"SAMA Rulebook - {sector_label}", f"Total documents: {len(documents)}",
        f"Categories: {len(by_category)}", f"With PDF: {with_pdf}",
    ])
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True, size=11, color="1F3864")
    ws.row_dimensions[1].height = 18
    ws.append([])

    ws.append(HEADERS)
    hdr_row = ws.max_row
    for cell in ws[hdr_row]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[hdr_row].height = 28

    pdf_col = HEADERS.index("PDF Link") + 1
    url_col = HEADERS.index("Document URL") + 1
    html_col = HEADERS.index("HTML File") + 1
    has_pdf_col = HEADERS.index("Has PDF") + 1

    html_dir = OUTPUT_DIR / "html" / re.sub(r'[\\/*?:"<>| ]', "_", sector_label)
    for i, doc in enumerate(documents, 1):
        html_file = _write_html_file(doc, i, html_dir=html_dir)
        row_data = _doc_to_row(doc, i, html_file, sector_label=sector_label)
        ws.append(row_data)
        row_idx = ws.max_row
        has_pdf = row_data[has_pdf_col - 1] == "YES"
        fill = ROW_FILL if (i % 2 == 0) else ALT_ROW_FILL
        if not has_pdf:
            fill = NO_PDF_FILL
        for cell in ws[row_idx]:
            cell.fill = fill
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

        pdf_cell = ws.cell(row=row_idx, column=pdf_col)
        if pdf_cell.value:
            pdf_cell.hyperlink = pdf_cell.value
            pdf_cell.font = LINK_FONT

        url_cell = ws.cell(row=row_idx, column=url_col)
        if url_cell.value:
            url_cell.hyperlink = url_cell.value
            url_cell.font = LINK_FONT

        html_cell = ws.cell(row=row_idx, column=html_col)
        if html_cell.value:
            html_cell.hyperlink = f"file:///{Path(html_cell.value).resolve().as_posix()}"
            html_cell.font = LINK_FONT

        ws.row_dimensions[row_idx].height = 15

    col_widths = {
        1: 6, 2: 7, 3: 55,
        4: 20, 5: 20, 6: 30, 7: 30, 8: 30, 9: 30, 10: 30, 11: 30,
        12: 45, 13: 18, 14: 16, 15: 16, 16: 8,
        17: 22, 18: 12, 19: 9, 20: 55, 21: 55,
        22: 12, 23: 45, 24: 70,
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = f"D{hdr_row + 1}"
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(HEADERS))}{hdr_row}"

    # Summary sheet: one row per category
    summary = wb.create_sheet("Category Summary")
    summary.append(["Category", "Document Count"])
    for cell in summary[1]:
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
    for cat, count in sorted(by_category.items(), key=lambda kv: -kv[1]):
        summary.append([cat, count])
    summary.column_dimensions["A"].width = 55
    summary.column_dimensions["B"].width = 16

    wb.save(str(out_path))
    return out_path


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="sama_finance_sector_requests.json")
    parser.add_argument("--output", default=str(EXCEL_DIR / "SAMA_Finance_Sector.xlsx"))
    args = parser.parse_args()

    documents = json.loads(Path(args.input).read_text(encoding="utf-8"))
    print(f"Loaded {len(documents)} documents from {args.input}")

    out_path = build_excel(documents, Path(args.output), sector_label="Finance Sector")
    print(f"Saved Excel report -> {out_path.resolve()}")
    print(f"Full per-document HTML files -> {(OUTPUT_DIR / 'html' / 'Finance_Sector').resolve()}")


if __name__ == "__main__":
    main()
